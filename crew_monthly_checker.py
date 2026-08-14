# ==========================================
# crew_monthly_checker.py
# 버전: v2.0 (2026-08-14) — crew_check.js v26 과 룰 동기화
# - 파서 재작성: 부분합류 크루(기장셀 빈 행) 오인식 버그 수정
# - 레그 단위 정밀 판정: 실제 담당 구간의 인원만 위반 판정
# - 세이프티 명단 월별 자동 적용(SP_BY_MONTH), 조회일 기준 자동선택
# - 1000시간 요건(생지공항 NTG/DAT/NGB/HET) 추가
# - 국내선 판정 기준: 레그 하나라도 국내면 포함(기존 전체구간→변경)
# - 공항 최신화(NGB, HET), ALV/CLV 표기 제거(실사용 없음 확인됨)
# - 편조 테이블 자동탐색(tables[1] 폴백 유지)
# 기능: CMS 편조점검 월간 자동 조회 (규정위반/내부위반/참고 엑셀 저장)
# 문의: 승무계획팀
# ==========================================
import asyncio
import re
import calendar
from datetime import datetime
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import messagebox
import tkinter as tk
import os

CMS_URL  = "https://crew.eastarjet.com/cms/Admin/Schedule/CrewPairs/CrewPairList.php"
HEADLESS = False

# ==========================================
# 편조점검 설정 (crew_check.js v26 과 동일)
# ==========================================
CFG = {
    "A": {"YNT","DSN","DAT","CGO","NGB","TXN","CGQ","SHE","HRB","MDC","KOJ","KMJ","IZO","TKS","TAE","CXR","DYG","DLC","YNJ","HKG","BSZ","ALA","MFM"},
    "B": {"NTG","HET","NRT","OKA","TSA","DAD","FUK","AOJ","PUS"},
    "C": {"PVG","KIX","CTS","KUV","ICN","GMP","CJJ","BKK","CNX","TPE","PQC","CJU"},
    "cxrBan" : {"신윤식","정진우"},
    "dadBan" : {"장준욱"},
    "foAonly": {"김상겸"},
    "foABonly":{"신영근"},
    "qa"     : {"박지현","신현욱","박승훈","신준서"},
    "cp"     : {"황종식","성기중","이재환","이태우"},
    "gradeOverride": {},
    # NTG/DAT/NGB/HET 4개 중국공항(생지공항): CPT 1000시간 이상자만 운항 가능 (승무팀 제공, 매월 갱신)
    "hr1000Airports": {"NTG","DAT","NGB","HET"},
    "hr1000": {"안선범","박상준","한상일","김도현","윤영규","류창상","조운영","이호성","신준서","김준식",
               "조웅진","신건수","박승훈","이상엽","김철균","김성엽","오병우","김경표","정진우","김우태",
               "김택의","사재철","김영준","오승민","정동일","정헌호","김병준","임승건","김범주","박한성",
               "김주성","김정희","김진욱","이유호","김치혁","여석윤","박승찬","라대영","정동수","박병구",
               "김현모","김대우","김병선","조재신","안태건","류재환","김상겸","김유진","이홍래","박태환",
               "김경태","이재환","이애릭","박기현","김국","신기철","문창환","유창욱","김의택","조준범",
               "최홍장","한가람","유영수","이마이클","권상준","이태우","이병주","임채홍","박상훈","신현욱",
               "백종혁","윤동희","이흥국","양세훈","정병국","김영채","류형년","노강철","김대연","허승혁",
               "신윤식","송필영","김윤태","문명성","황종식","김효진","박지현","유동윤","성기중","김재훈",
               "이민영","남준현","배대익","유영우","김병주","김찬수","주재도","손동현","박재일","이동화",
               "이준민","이용승","이경혁","이일주","장준욱","신영근","안영환"},
}

# ── 월별 세이프티(FO) 불가/예외 명단 ──
# 조회 대상 날짜(YYYY-MM) 기준으로 자동 선택
SP_BY_MONTH = {
    "2026-07": {"ban": {"김창중","이주화","양병모","엄태국","김우영","최은총","장재봉","이창민",
                         "이한솔","정종성","김공주","김총화","김재영","이웅배","김민재","한다영","최도현"},
                "ok": {"엄태국","양병모"}},
    "2026-08": {"ban": {"김창중","이주화","김우영","최은총","장재봉","이창민","이한솔","정종성",
                         "김공주","김총화","김재영","이웅배","김민재","최도현"},
                "ok": set()},
    "2026-09": {"ban": {"김창중","최은총","장재봉","이창민","이한솔","정종성","김공주","김총화",
                         "김재영","이웅배","김민재","최도현","이재현","윤동건","한건희","박신우",
                         "배민수","진석준"},
                "ok": set()},
}

def get_sp_lists(ym):
    """조회월(YYYY-MM) 기준 SP 명단 선택. 없으면 가장 가까운 이전월, 그마저 없으면 가장 이른월."""
    keys = sorted(SP_BY_MONTH.keys())
    if ym in SP_BY_MONTH:
        key = ym
    else:
        earlier = [k for k in keys if k <= ym]
        key = earlier[-1] if earlier else keys[0]
    return SP_BY_MONTH[key]["ban"], SP_BY_MONTH[key]["ok"], key

NAME_RE  = re.compile(r'[가-힣]{2,5}(?:[ABCX](?:LV)?)?')
GRADE_RE = re.compile(r'^[가-힣]{2,5}([ABCX])(?:LV)?')
LV_RE    = re.compile(r'^[가-힣]{2,5}[ABCX](LV)$')

def get_name(s):
    return re.sub(r'[ABCX](?:LV)?.*$', '', s)

def get_site_grade(s):
    m = GRADE_RE.match(s)
    return m.group(1) if m else ''

def get_grade(s):
    n = get_name(s)
    if n in CFG["gradeOverride"]:
        return CFG["gradeOverride"][n]
    return get_site_grade(s)

def get_lv(s):
    m = LV_RE.match(s)
    return m.group(1) if m else ''

def is_junk(s):
    return bool(re.match(r'^\d{1,2}/\d{1,2}', s)) or '편조' in s or '점검' in s or len(s) == 0

# 국내선 판별: 레그 하나라도 국내면 국내선으로 간주 (crew_check.js v19 이후 기준)
KR = {"ICN", "GMP", "CJU", "CJJ", "KUV", "PUS", "TAE"}
def is_dom(rt):
    p = str(rt or '').split('/')
    return len(p) > 1 and p[0] in KR and p[1] in KR

def pick_table(tables):
    """편조 테이블 자동 탐색: 노선 패턴(XXX/XXX)이 가장 많은 테이블 선택.
    CMS 구조가 바뀌어 못 찾으면 기존 방식(tables[1])으로 폴백."""
    best, best_score = None, 0
    route_re = re.compile(r'[A-Z]{3,4}/[A-Z]{3,4}')
    for t in tables:
        txt = t.get_text(' ', strip=True)
        score = len(route_re.findall(txt))
        if score > best_score:
            best_score, best = score, t
    if best_score >= 3:
        return best
    return tables[1] if len(tables) > 1 else None

def parse_table(html):
    """행(tr) 단위로 셀 파싱. 기장셀(cells[0]) 유무를 hasCap 으로 보존."""
    soup = BeautifulSoup(html, 'html.parser')
    tables = soup.find_all('table')
    tb = pick_table(tables)
    if tb is None:
        return []
    rows = []
    for tr in tb.find_all('tr'):
        cells = tr.find_all('td')
        if not cells or len(cells) < 3:
            t = tr.get_text(' ', strip=True)
            if t:
                rows.append({'line': t, 'hasCap': True})
            continue

        def cell_txt(c):
            txt = re.sub(r' {2,}', ' ', c.get_text(' ', strip=True))
            # CMS HTML: 이름과 등급 사이 공백 제거 ("정병국 A" → "정병국A")
            txt = re.sub(r'([가-힣]) ([ABCX](?:LV)?)\b', r'\1\2', txt)
            return txt

        cap_cell   = cell_txt(cells[0])
        fo_cell    = cell_txt(cells[1])
        extra_cell = cell_txt(cells[2]) if len(cells) > 2 else ''

        caps   = NAME_RE.findall(cap_cell)
        fos    = NAME_RE.findall(fo_cell)
        extras = NAME_RE.findall(extra_cell)

        rest = ' '.join(cell_txt(cells[i]) for i in range(3, len(cells)))

        ordered = []
        for k in range(max(len(caps), len(fos))):
            if k < len(caps): ordered.append(caps[k])
            if k < len(fos):  ordered.append(fos[k])
        ordered += extras

        line = (' '.join(ordered) + ' ' + rest).strip()
        if line:
            rows.append({'line': line, 'hasCap': len(caps) > 0})
    return rows

TOKEN_RE = re.compile(
    r'(\d{2}:\d{2})'
    r'|([A-Z]{3,4}/[A-Z]{3,4})'
    r'|(\d{3,4})(?![\d:])'
    r'|([가-힣]{2,5}(?:[ABCX](?:LV)?)?)'
)

def make_legs(flights, fo, extra):
    return [{'fl': f['fl'], 'rt': f['rt'], 'fo': fo or '', 'extra': list(extra or [])} for f in flights]

def parse_blocks(row_objs):
    """행(row) 단위로 block 생성. 기장셀 없는 행은 편명 겹치는 편조에 레그 단위로 정확히 병합."""
    # LV 단독 라인은 앞 행 마지막 이름에 병합
    merged = []
    for o in row_objs:
        L = o['line']
        if re.match(r'^LV$', L) and merged and re.search(r'[가-힣]{2,5}[ABCX]?$', merged[-1]['line']):
            merged[-1]['line'] += L
        else:
            merged.append({'line': L, 'hasCap': o['hasCap']})
    clean = [o for o in merged if not is_junk(o['line'])]

    # 각 행(line)을 하나의 block으로 파싱 (line=block 1:1), hasCap 보존
    blocks = []
    for o in clean:
        typed = []
        for m in TOKEN_RE.finditer(o['line']):
            if m.group(1): typed.append(('time', m.group(1)))
            elif m.group(2): typed.append(('route', m.group(2)))
            elif m.group(3): typed.append(('flight', m.group(3)))
            elif m.group(4): typed.append(('name', m.group(4)))
        names, flights, i, N = [], [], 0, len(typed)
        while i < N and typed[i][0] == 'name':
            names.append(typed[i][1]); i += 1
        while i < N:
            if typed[i][0] == 'flight':
                f = typed[i][1]; i += 1
                if i < N and typed[i][0] == 'route':
                    r = typed[i][1]; i += 1
                    while i < N and typed[i][0] == 'time': i += 1
                    flights.append({'fl': f, 'rt': r})
            else:
                i += 1
        if names or flights:
            blocks.append({'names': names, 'flights': flights, 'hasCap': o['hasCap']})

    # 편명만 있고 이름 없는 block은 앞 block에 편명 흡수 (연속 편명 대응)
    blk2 = []
    for b in blocks:
        if not b['names'] and b['flights'] and blk2:
            blk2[-1]['flights'].extend(b['flights'])
        elif b['names']:
            blk2.append(b)

    # 분류: 기장셀 있는 행 → main / 기장셀 없는 행 → pending(부분합류 후보)
    mains, pending, solos = [], [], []
    for b in blk2:
        if b['hasCap']:
            fo = b['names'][1] if len(b['names']) > 1 else ''
            extra = b['names'][2:]
            mains.append({'cap': b['names'][0], 'fo': fo, 'extra': extra,
                           'flights': b['flights'], 'legs': make_legs(b['flights'], fo, extra)})
        else:
            pending.append(b)

    # 부분합류 병합: 기장 없는 행을 편명 겹치는 편조에 레그 단위로 정확히 붙임
    for p in pending:
        all_xor = all(get_grade(n) in ('X', '') for n in p['names'])
        if all_xor:
            p['asSolo'] = True
            solos.append(p)
            continue
        pfl = {f['fl'] for f in p['flights']}
        target, best = None, 0
        for m in mains:
            mfl = {f['fl'] for f in m['flights']}
            cnt = len(pfl & mfl)
            if cnt > best:
                best, target = cnt, m
        if target and best > 0:
            p_fo = p['names'][0]
            p_extra = p['names'][1:]
            for leg in target['legs']:
                if leg['fl'] in pfl:
                    leg['fo'] = p_fo
                    leg['extra'] = p_extra
            # 대표 표시값(패널/엑셀 상단 cap/fo 요약용)
            for nm in p['names']:
                if not target['fo']:
                    target['fo'] = nm
                else:
                    target['extra'].append(nm)
        elif len(p['names']) >= 2:
            fo = p['names'][1]
            extra = p['names'][2:]
            mains.append({'cap': p['names'][0], 'fo': fo, 'extra': extra,
                           'flights': p['flights'], 'legs': make_legs(p['flights'], fo, extra)})
        else:
            p['asSolo'] = True
            solos.append(p)

    # 4인편성 split
    result = []
    for m in mains:
        all_n = [m['cap'], m['fo']] + m['extra']
        graded  = [n for n in all_n if get_grade(n) in ('A', 'B', 'C')]
        nograde = [n for n in all_n if get_grade(n) not in ('A', 'B', 'C', 'X')]
        if len(graded) >= 4:
            result.append({'cap': graded[0], 'fo': graded[1], 'extra': nograde,
                            'flights': m['flights'], 'legs': make_legs(m['flights'], graded[1], nograde)})
            result.append({'cap': graded[2], 'fo': graded[3], 'extra': [],
                            'flights': m['flights'], 'legs': make_legs(m['flights'], graded[3], [])})
        else:
            result.append(m)

    for s in solos:
        if s.get('asSolo'):
            result.append({'isSolo': True, 'names': s['names'], 'flights': s['flights']})
    return result

def group_legs(legs):
    """fo/extra 조합이 같은 연속 레그를 하나의 그룹으로 묶음 (실제 담당 구간 기준)."""
    groups = []
    for leg in (legs or []):
        key = leg['fo'] + '|' + ','.join(leg['extra'])
        if groups and groups[-1]['key'] == key:
            groups[-1]['flights'].append({'fl': leg['fl'], 'rt': leg['rt']})
        else:
            groups.append({'key': key, 'fo': leg['fo'], 'extra': leg['extra'],
                            'flights': [{'fl': leg['fl'], 'rt': leg['rt']}]})
    return groups

def check(blocks, sp_ban, sp_ok):
    violations, internalV = [], []
    seen = {'cc': set(), 'cf': set(), 'aa': set()}
    fl_set = set()

    for b in blocks:
        if b.get('isSolo'):
            fl_set.update(f['fl'] for f in b['flights'])
            fls0 = '/'.join(f['fl'] for f in b['flights'])
            dom0 = bool(b['flights']) and any(is_dom(f['rt']) for f in b['flights'])
            for n in b.get('names', []):
                g = get_grade(n)
                label = 'DH/훈련' if g == 'X' else '추가 탑승'
                internalV.append({'type': '참고', 'note': True, 'detail': label,
                                   'fl': fls0, 'pair': get_name(n), 'dom': dom0})
            continue

        cap_n, cap_g = get_name(b['cap']), get_grade(b['cap'])
        fo_disp = get_name(b['fo']) if b['fo'] else ''
        fls = '/'.join(f['fl'] for f in b['flights'])
        fl_set.update(f['fl'] for f in b['flights'])
        cur_dom = bool(b['flights']) and any(is_dom(f['rt']) for f in b['flights'])

        # 사람 속성(사이트등급 갱신/LV/심사관)은 편조 대표 인원 기준 1회만 표시
        for raw in [b['cap'], b['fo']] + list(b.get('extra', [])):
            if not raw:
                continue
            nm, sg = get_name(raw), get_site_grade(raw)
            if nm in CFG['gradeOverride'] and sg == CFG['gradeOverride'][nm]:
                internalV.append({'type': '참고', 'note': True,
                                   'detail': '✅사이트 등급 갱신 확인(오버라이드 해제 가능)',
                                   'fl': fls, 'pair': f"{nm}({sg})", 'dom': cur_dom})
            lv = get_lv(raw)
            if lv:
                internalV.append({'type': '참고', 'note': True, 'detail': '🌫️LV 저시정 제한',
                                   'fl': fls, 'pair': f"{get_name(raw)} ({get_grade(raw)}{lv})", 'dom': cur_dom})
        if cap_n in CFG['qa']:
            internalV.append({'type':'참고','note':True,'detail':'ℹ️품질심사관','fl':fls,'pair':cap_n,'dom':cur_dom})
        if fo_disp and get_name(b['fo']) in CFG['qa']:
            internalV.append({'type':'참고','note':True,'detail':'ℹ️품질심사관','fl':fls,'pair':get_name(b['fo']),'dom':cur_dom})
        if cap_n in CFG['cp']:
            internalV.append({'type':'참고','note':True,'detail':'ℹ️노선심사관','fl':fls,'pair':cap_n,'dom':cur_dom})
        if fo_disp and get_name(b['fo']) in CFG['cp']:
            internalV.append({'type':'참고','note':True,'detail':'ℹ️노선심사관','fl':fls,'pair':get_name(b['fo']),'dom':cur_dom})

        # ── 규정/세이프티 판정은 실제 담당 구간(레그 그룹) 단위로 정확히 처리 ──
        src_legs = b.get('legs') or [{'fl': f['fl'], 'rt': f['rt'], 'fo': b['fo'], 'extra': b.get('extra', [])} for f in b['flights']]
        groups = group_legs(src_legs)

        for grp in groups:
            grp_fo = grp['fo'] or ''
            grp_extra = grp['extra'] or []
            grp_fo_n, grp_fo_g = get_name(grp_fo), get_grade(grp_fo)
            grp_fo_eff = 'SKIP' if grp_fo_g == 'X' else (grp_fo_g if grp_fo_g else '')
            grp_fls = '/'.join(f['fl'] for f in grp['flights'])
            pair = f"{b['cap']}/{grp_fo or '-'}"
            grp_dom = any(is_dom(f['rt']) for f in grp['flights'])

            # 세이프티
            has_trainee = (cap_g in ('', 'X')) or (grp_fo_g in ('', 'X')) or any(get_grade(e) in ('', 'X') for e in grp_extra)
            if has_trainee:
                for rw in [b['cap'], grp_fo] + list(grp_extra):
                    if not rw:
                        continue
                    nm, g = get_name(rw), get_grade(rw)
                    if g in ('', 'X'):
                        continue
                    if nm not in sp_ban:
                        continue
                    if nm in sp_ok:
                        internalV.append({'type':'참고','note':True,'detail':'ℹ️SP 예외자(세이프티 가능)',
                                           'fl':grp_fls,'pair':f"{nm} - 훈련/관숙 동승, 세이프티 가능",'dom':grp_dom})
                    else:
                        internalV.append({'type':'내부위반','detail':'세이프티 불가자 + 훈련/관숙 동승 (확인 필요)',
                                           'fl':grp_fls,'pair':f"{nm} / {pair}",'dom':grp_dom})

            for flt in grp['flights']:
                org, dst = (flt['rt'].split('/') if '/' in flt['rt'] else (flt['rt'], ''))

                if cap_g == 'C':
                    ok = grp_fo_eff == 'A'
                    obs = None
                    if not ok and grp_fo_eff in ('SKIP', ''):
                        obs = next((e for e in grp_extra if get_grade(e) == 'A'), None)
                        if obs: ok = True
                    ck = f"cc|{pair}"
                    if ck not in seen['cc']:
                        seen['cc'].add(ck)
                    if not ok:
                        msg = 'C기장 관숙 편성 위반(FO A 동승 필요)' if grp_fo_eff in ('SKIP', '') else 'C기장 페어링 위반'
                        violations.append({'type':'규정위반','detail':msg,'fl':flt['fl'],'pair':pair,'ap':'','dom':grp_dom})
                    for ap in (org, dst):
                        if ap in CFG['B']:
                            violations.append({'type':'규정위반','detail':'B공항 C기장 위반','fl':flt['fl'],'pair':pair,'ap':ap,'dom':grp_dom})
                        if ap and ap not in CFG['A'] and ap not in CFG['B'] and ap not in CFG['C']:
                            violations.append({'type':'규정위반','detail':'C기장 분류외 공항 위반','fl':flt['fl'],'pair':pair,'ap':ap,'dom':grp_dom})

                if grp_fo_eff == 'C':
                    ok2 = cap_g == 'A'
                    if not ok2:
                        violations.append({'type':'규정위반','detail':'C부기장 페어링 위반','fl':flt['fl'],'pair':pair,'ap':'','dom':grp_dom})
                    for ap in (org, dst):
                        if ap and ap not in CFG['A'] and ap not in CFG['B'] and ap not in CFG['C']:
                            violations.append({'type':'규정위반','detail':'C부기장 분류외 공항 위반','fl':flt['fl'],'pair':pair,'ap':ap,'dom':grp_dom})

                for ap in (org, dst):
                    if ap in CFG['A']:
                        cok = cap_g == 'A'
                        fok = grp_fo_eff in ('A', 'SKIP')
                        if not cok:
                            violations.append({'type':'규정위반','detail':'A공항 기장 등급 위반','fl':flt['fl'],'pair':pair,'ap':ap,'dom':grp_dom})
                        if not fok:
                            violations.append({'type':'규정위반','detail':'A공항 부기장 등급 위반','fl':flt['fl'],'pair':pair,'ap':ap,'dom':grp_dom})
                    if ap == 'CXR' and cap_n in CFG['cxrBan']:
                        internalV.append({'type':'내부위반','detail':'CXR 금지','fl':flt['fl'],'pair':b['cap'],'ap':'','dom':grp_dom})
                    if ap == 'DAD' and cap_n in CFG['dadBan']:
                        internalV.append({'type':'내부위반','detail':'DAD 금지','fl':flt['fl'],'pair':b['cap'],'ap':'','dom':grp_dom})
                    if ap in CFG['hr1000Airports']:
                        if cap_n not in CFG['hr1000']:
                            violations.append({'type':'규정위반','detail':'1000시간 미만 기장 운항 (생지공항 전용요건)',
                                                'fl':flt['fl'],'pair':b['cap'],'ap':ap,'dom':grp_dom})
                        else:
                            internalV.append({'type':'참고','note':True,'detail':'🛫생지공항(1000hr↑) 정상',
                                               'fl':grp_fls,'pair':f"{b['cap']} ({ap})",'dom':grp_dom})

                if cap_n in CFG['foAonly']:
                    if grp_fo_eff != 'A':
                        internalV.append({'type':'내부위반','detail':f'{cap_n} FO제한위반','fl':flt['fl'],'pair':grp_fo,'ap':'','dom':grp_dom})
                if cap_n in CFG['foABonly']:
                    if grp_fo_eff not in ('A', 'B'):
                        internalV.append({'type':'내부위반','detail':f'{cap_n} FO제한위반','fl':flt['fl'],'pair':grp_fo,'ap':'','dom':grp_dom})

            if grp_fo_g == 'X':
                internalV.append({'type':'참고','note':True,'detail':'DH/훈련','fl':grp_fls,'pair':grp_fo_n,'dom':grp_dom})
            if grp_extra:
                ng = [e for e in grp_extra if get_grade(e) == '' and re.match(r'^[가-힣]{2,4}$', e)]
                g2 = [e for e in grp_extra if get_grade(e) in ('A', 'B', 'C')]
                gx = [e for e in grp_extra if get_grade(e) == 'X']
                if ng:
                    internalV.append({'type':'참고','note':True,'detail':'추가탑승','fl':grp_fls,'pair':','.join(ng),'dom':grp_dom})
                if g2:
                    internalV.append({'type':'참고','note':True,'detail':'추가탑승','fl':grp_fls,'pair':','.join(g2),'dom':grp_dom})
                if gx:
                    internalV.append({'type':'참고','note':True,'detail':'DH/훈련','fl':grp_fls,
                                       'pair':','.join(get_name(e) for e in gx),'dom':grp_dom})

    return violations, internalV

def save_excel(all_results, year, month):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년{month:02d}월_위반사항"

    hdr_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    v_fill    = PatternFill('solid', start_color='3A1E1E')
    i_fill    = PatternFill('solid', start_color='3A2E1E')
    n_fill    = PatternFill('solid', start_color='1E2A3A')
    center = Alignment(horizontal='center', vertical='center')
    wrap   = Alignment(wrap_text=True, vertical='center')

    headers = ['날짜', '구분', '편명', '위반유형', '페어링', '공항', '국내/국제']
    widths  = [12, 8, 12, 32, 28, 8, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, col, h)
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, center
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 20

    row_idx = 2
    total_v = total_i = 0

    for date_str, violations, internalV in all_results:
        for v in violations:
            ws.cell(row_idx, 1, date_str).alignment = center
            ws.cell(row_idx, 2, '🚨규정위반').alignment = center
            ws.cell(row_idx, 3, v['fl']).alignment = center
            ws.cell(row_idx, 4, v['detail']).alignment = wrap
            ws.cell(row_idx, 5, v['pair']).alignment = wrap
            ws.cell(row_idx, 6, v.get('ap','')).alignment = center
            ws.cell(row_idx, 7, '국내선' if v.get('dom') else '국제선').alignment = center
            for col in range(1, 8):
                ws.cell(row_idx, col).fill = v_fill
                ws.cell(row_idx, col).font = Font(name='맑은 고딕', size=10)
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1
            total_v += 1

        for v in internalV:
            is_note = v.get('note', False)
            fill = n_fill if is_note else i_fill
            ws.cell(row_idx, 1, date_str).alignment = center
            ws.cell(row_idx, 2, 'ℹ️참고' if is_note else '⚠️내부위반').alignment = center
            ws.cell(row_idx, 3, v['fl']).alignment = center
            ws.cell(row_idx, 4, v['detail']).alignment = wrap
            ws.cell(row_idx, 5, v['pair']).alignment = wrap
            ws.cell(row_idx, 6, v.get('ap','')).alignment = center
            ws.cell(row_idx, 7, '국내선' if v.get('dom') else '국제선').alignment = center
            for col in range(1, 8):
                ws.cell(row_idx, col).fill = fill
                ws.cell(row_idx, col).font = Font(name='맑은 고딕', size=10)
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1
            if not is_note:
                total_i += 1

    if row_idx == 2:
        ws.cell(2, 1, '✅ 위반사항 없음')
        ws.cell(2, 1).font = Font(name='맑은 고딕', bold=True, color='4ADE80', size=11)
    else:
        ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{row_idx - 1}"

    ws2 = wb.create_sheet('요약')
    ws2['A1'] = f"{year}년 {month:02d}월 편조점검 결과"
    ws2['A1'].font = Font(name='맑은 고딕', bold=True, size=13)
    ws2['A3'] = '조회 일수'
    ws2['B3'] = len(all_results)
    ws2['A4'] = '규정위반 건수'
    ws2['B4'] = total_v
    ws2['B4'].font = Font(color='FF6B6B' if total_v else '4ADE80', bold=True)
    ws2['A5'] = '내부위반 건수'
    ws2['B5'] = total_i
    ws2['B5'].font = Font(color='FFD166' if total_i else '4ADE80', bold=True)
    ws2['A6'] = '생성일시'
    ws2['B6'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 20
    default_font = Font(name='맑은 고딕', size=11)
    for row in ws2.iter_rows(min_row=3, max_row=6):
        for cell in row:
            if not cell.font or not cell.font.name:
                cell.font = default_font

    out_path = os.path.join(
        os.path.expanduser('~'), 'Desktop',
        f"편조점검_{year}{month:02d}.xlsx"
    )
    wb.save(out_path)
    return out_path, total_v, total_i

def get_target_month():
    """이번 달 / 다음 달 선택 팝업"""
    root = tk.Tk()
    root.withdraw()
    today = datetime.now()
    this_y, this_m = today.year, today.month
    if this_m == 12:
        next_y, next_m = this_y + 1, 1
    else:
        next_y, next_m = this_y, this_m + 1

    answer = messagebox.askquestion(
        "조회 월 선택",
        f"조회할 월을 선택하세요.\n\n"
        f" [예] 다음 달 ({next_y}년 {next_m:02d}월, 1일~말일)\n"
        f" [아니오] 이번 달 ({this_y}년 {this_m:02d}월, 오늘~말일)",
        icon="question"
    )
    root.destroy()

    if answer == "yes":
        last = calendar.monthrange(next_y, next_m)[1]
        return next_y, next_m, 1, last
    else:
        last = calendar.monthrange(this_y, this_m)[1]
        return this_y, this_m, today.day, last

async def main():
    print('='*50)
    print('✈  편조점검 월간 자동 조회 v2.0')
    print('    (2026-08-14) | 문의: 승무계획팀')
    print('='*50)

    year, month, start_day, last_day = get_target_month()
    print(f"\n조회 대상: {year}년 {month:02d}월 ({start_day}일 ~ {last_day}일)\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS, args=[
            '--disable-blink-features=AutomationControlled',
            '--no-sandbox', '--window-size=1280,800'
        ])
        context = await browser.new_context(
            locale='ko-KR', timezone_id='Asia/Seoul',
            viewport={'width': 1280, 'height': 800}
        )
        page = await context.new_page()

        await page.goto(CMS_URL, wait_until='domcontentloaded', timeout=20000)
        print('브라우저가 열렸습니다.')
        print('CMS에 로그인 후 엔터를 눌러주세요...')
        await asyncio.get_event_loop().run_in_executor(None, input, '  [로그인 완료 후 엔터] ')

        all_results = []

        for day in range(start_day, last_day + 1):
            date_str = f"{year}/{month:02d}/{day:02d}"
            ym = f"{year}-{month:02d}"
            sp_ban, sp_ok, sp_key = get_sp_lists(ym)
            url = f"{CMS_URL}?d={year}-{month:02d}-{day:02d}&s=true&e=true&a=true"
            print(f"[{day:02d}/{last_day}] {date_str} 조회 중...", end=' ', flush=True)

            try:
                await page.goto(url, wait_until='networkidle', timeout=30000)
                await page.wait_for_timeout(2500)

                body = await page.inner_text('body')
                if 'SESSION EXPIRED' in body or 'logout' in page.url:
                    print('⚠️ 세션 만료 → 재로그인 후 엔터')
                    await asyncio.get_event_loop().run_in_executor(None, input, '  [재로그인 후 엔터] ')
                    await page.goto(url, wait_until='networkidle', timeout=30000)
                    await page.wait_for_timeout(2500)

                html = await page.content()
                row_objs = parse_table(html)

                if not row_objs:
                    print('데이터 없음 (스킵)')
                    continue

                blocks = parse_blocks(row_objs)
                violations, internalV = check(blocks, sp_ban, sp_ok)

                if violations or internalV:
                    all_results.append((date_str, violations, internalV))
                    n_int = sum(1 for v in internalV if not v.get('note'))
                    n_note = len(internalV) - n_int
                    note_txt = f" / ℹ️ 참고 {n_note}건" if n_note else ''
                    print(f"🚨 규정위반 {len(violations)}건 / ⚠️ 내부위반 {n_int}건{note_txt}")
                    for v in violations:
                        ap = f" ({v['ap']})" if v.get('ap') else ''
                        print(f"    🚨 [{v['fl']}] {v['detail']}{ap} | {v['pair']}")
                    for v in internalV:
                        icon = 'ℹ️ ' if v.get('note') else '⚠️ '
                        print(f"    {icon} [{v['fl']}] {v['detail']} | {v['pair']}")
                else:
                    print('✅ 이상없음')

                await page.wait_for_timeout(1200)

            except PWTimeout:
                print('⏱️ 타임아웃 (스킵)')
            except Exception as e:
                print(f'💥 오류: {e}')

        await browser.close()

    if not all_results:
        print(f'\n✅ {year}년 {month:02d}월 전체 위반사항 없음')
        input('\n엔터 누르면 종료...')
        return

    out_path, total_v, total_i = save_excel(all_results, year, month)
    print(f'\n{"="*50}')
    print(f'✅ 조회 완료: {len(all_results)}일 위반 발생 ({start_day}일~{last_day}일)')
    print(f'   🚨 규정위반: {total_v}건')
    print(f'   ⚠️  내부위반: {total_i}건')
    print(f'\n저장 완료: {out_path}')
    input('\n엔터 누르면 종료...')

if __name__ == '__main__':
    asyncio.run(main())
